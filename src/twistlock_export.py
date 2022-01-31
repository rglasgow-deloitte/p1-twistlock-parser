import enum
from fileinput import filename
import os
import json
from re import A
import openpyxl


class TwistlockExport:
    def __init__(self, file_queue):
        self.cache = {}
        self.file_queue = file_queue
        self.column_names = [
            "cve",
            "link",
            "packageName",
            "packageVersion",
            "severity",
            "status",
        ]

    def __parse_twistlock_json(self, file_name):
        # open the json file found in the directory 'clean'
        with open(f"clean/{file_name}", "r") as f:
            # read the json file
            data = f.read()
            # parse the json file
            data = json.loads(data)
            # return the data
            return data

    def __add_to_cache(self, data, file_name=None):
        """
        Adds data to the cache while being parsed
        """
        # in the data, grab the cve
        cve = data["cve"]

        # if the cve is not in the cache
        if cve not in self.cache:
            # unravel the data
            new_obj = data
            new_obj["file_name"] = f"{file_name}"
            # add the data to the cache
            self.cache[cve] = new_obj
        else:
            # update the data.file_name to include the file name
            new_obj = self.cache[cve]

            # verify the name of the file is not already in the obj
            if new_obj["file_name"].find(file_name) == -1:
                # add the file name to the obj
                new_obj["file_name"] += f", {file_name}"

            # add the data to the cache
            self.cache[cve] = new_obj

    def __parse_cve(self, sheet, data, column_names, file_name):
        # iterate through each column of the cve data and add it to the sheet
        for index in range(len(data)):
            data_to_cache = {}

            # use each column name as the key
            for column_index in range(len(column_names)):
                # add the data to the cache
                data_to_cache[column_names[column_index]] = data[index][
                    column_names[column_index]
                ]

                # add the data to the sheet
                sheet.cell(row=index + 2, column=column_index + 1).value = data[index][
                    column_names[column_index]
                ]

            # add the data to the cache
            self.__add_to_cache(data_to_cache, file_name)

    def __write_cve_to_sheet(self, sheet, file_name):
        column_names = [
            "cve",
            "link",
            "packageName",
            "packageVersion",
            "severity",
            "status",
        ]
        # iterate through each column name
        for index in range(len(column_names)):
            # use index as the row
            sheet.cell(row=1, column=index + 1).value = column_names[index]

        # grab the cve data from the json file
        cve_data = self.__parse_twistlock_json(file_name)["cve"]
        self.__parse_cve(sheet, cve_data, column_names, file_name)

    def __write_compliance_to_sheet(self, sheet):

        pass

    def __write_cache_to_sheet(self, sheet, file_name):
        # create a new sheet for the cache
        cache_columns = self.column_names + ["file_name"]

        # get one row of data from the cache
        keys = self.cache.keys()
        first_row = {}
        key_index = 0
        for key in keys:
            if key_index > 0:
                break
            first_row = self.cache[key]
            key_index += 1

        # iterate through each first row key
        count = 0
        for key in first_row.keys():
            # add the key to the cache sheet
            sheet.cell(row=1, column=count + 1).value = key
            count += 1

        # iterate through each cve in the cache
        for index in range(len(self.cache)):
            # grab the cve
            cve = list(self.cache.keys())[index]
            # grab the data
            data = self.cache[cve]

            # iterate through each key in the data
            for key in data:
                # add the data to the cache sheet
                sheet.cell(
                    row=index + 2, column=cache_columns.index(key) + 1
                ).value = data[key]

    def export_to_excel(self):
        """
        Prepares data for excel
        """
        # create a new directory only if it doesn't exist 'excel'
        if not os.path.exists("excel"):
            os.mkdir("excel")

        # create a new workbook
        workbook = openpyxl.Workbook()

        # for each file in the queue
        for file in self.file_queue:
            # create a new sheet
            current_sheet = workbook.create_sheet(file)
            self.__write_cve_to_sheet(current_sheet, file)

        cache_sheet = workbook.create_sheet("vulnerability_cross_reference")
        self.__write_cache_to_sheet(cache_sheet, file)

        # write compliance to sheet
        compliance_sheet = workbook.create_sheet("compliance")
        self.__write_compliance_to_sheet(compliance_sheet)

        # save the workbook
        workbook.save("excel/twistlock_export.xlsx")
